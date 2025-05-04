import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_excel_template():
    # สร้างไฟล์ Excel ใหม่
    wb = openpyxl.Workbook()
    
    # สร้างชีท Projects
    projects_sheet = wb.active
    projects_sheet.title = "Projects"
    
    # กำหนดหัวข้อคอลัมน์และคำอธิบายสำหรับชีท Projects
    projects_columns = [
        {"header": "id", "description": "รหัสโครงการ (ต้องไม่ซ้ำกัน)", "example": "PRJ001", "note": "ต้องเป็นตัวอักษรหรือตัวเลข ไม่มีช่องว่าง"},
        {"header": "name", "description": "ชื่อโครงการ", "example": "โครงการพัฒนาคุณภาพการศึกษา", "note": "ควรระบุให้ชัดเจนและเข้าใจง่าย"},
        {"header": "responsiblePerson", "description": "ผู้รับผิดชอบโครงการ", "example": "นายสมชาย ใจดี", "note": "ระบุชื่อ-นามสกุลให้ครบถ้วน"},
        {"header": "budget", "description": "งบประมาณรวม (บาท)", "example": "150000", "note": "ต้องเป็นตัวเลขเท่านั้น ไม่มีเครื่องหมายคอมม่า"},
        {"header": "workGroup", "description": "กลุ่มงาน", "example": "academic", "note": "ต้องเลือกจาก: academic, budget, hr, general, other"},
        {"header": "startDate", "description": "วันที่เริ่มต้น", "example": "2023-01-01", "note": "ต้องเป็นรูปแบบ YYYY-MM-DD"},
        {"header": "endDate", "description": "วันที่สิ้นสุด", "example": "2023-12-31", "note": "ต้องเป็นรูปแบบ YYYY-MM-DD และต้องไม่น้อยกว่าวันที่เริ่มต้น"},
        {"header": "status", "description": "สถานะโครงการ", "example": "active", "note": "ต้องเลือกจาก: active, completed"},
        {"header": "description", "description": "รายละเอียดโครงการ", "example": "โครงการเพื่อพัฒนาคุณภาพการศึกษา...", "note": "ควรระบุวัตถุประสงค์และรายละเอียดที่สำคัญ"},
        {"header": "budgetCategories", "description": "หมวดงบประมาณ (JSON)", "example": '[{"category":"เงินอุดหนุนรายหัว","amount":100000},{"category":"เงินพัฒนาผู้เรียน","amount":50000}]', "note": "ต้องเป็น JSON array ที่ถูกต้อง จำนวนเงินรวมต้องเท่ากับงบประมาณรวม"}
    ]
    
    # สร้างชีท Transactions
    transactions_sheet = wb.create_sheet("Transactions")
    
    # กำหนดหัวข้อคอลัมน์และคำอธิบายสำหรับชีท Transactions
    transactions_columns = [
        {"header": "id", "description": "รหัสรายการ (ต้องไม่ซ้ำกัน)", "example": "TRX001", "note": "ต้องเป็นตัวอักษรหรือตัวเลข ไม่มีช่องว่าง"},
        {"header": "projectId", "description": "รหัสโครงการที่เกี่ยวข้อง", "example": "PRJ001", "note": "ต้องตรงกับรหัสโครงการที่มีอยู่ในระบบ"},
        {"header": "date", "description": "วันที่ทำรายการ", "example": "2023-02-15", "note": "ต้องเป็นรูปแบบ YYYY-MM-DD"},
        {"header": "description", "description": "รายละเอียดรายการ", "example": "จัดซื้อวัสดุการเรียน", "note": "ควรระบุรายละเอียดให้ชัดเจน"},
        {"header": "amount", "description": "จำนวนเงิน (บาท)", "example": "5000", "note": "ต้องเป็นตัวเลขเท่านั้น ไม่มีเครื่องหมายคอมม่า"},
        {"header": "budgetCategory", "description": "หมวดงบประมาณ", "example": "เงินอุดหนุนรายหัว", "note": "ต้องตรงกับหมวดที่กำหนดในโครงการ"},
        {"header": "note", "description": "หมายเหตุ", "example": "เบิกจ่ายตามใบเสร็จเลขที่ 12345", "note": "ไม่บังคับ กรณีมีเอกสารอ้างอิงควรระบุ"},
        {"header": "isTransfer", "description": "เป็นรายการโอนเงินหรือไม่", "example": "true", "note": "ต้องเป็น true หรือ false เท่านั้น"},
        {"header": "isTransferIn", "description": "เป็นรายการรับโอนหรือไม่", "example": "true", "note": "ต้องเป็น true หรือ false เท่านั้น (กรณี isTransfer เป็น true)"},
        {"header": "transferToProjectId", "description": "รหัสโครงการปลายทาง", "example": "PRJ002", "note": "ต้องระบุเมื่อ isTransfer เป็น true"},
        {"header": "transferToCategory", "description": "หมวดงบประมาณปลายทาง", "example": "เงินพัฒนาผู้เรียน", "note": "ต้องระบุเมื่อ isTransfer เป็น true"},
        {"header": "transferFromProjectId", "description": "รหัสโครงการต้นทาง", "example": "PRJ001", "note": "ต้องระบุเมื่อ isTransfer เป็น true"},
        {"header": "transferFromCategory", "description": "หมวดงบประมาณต้นทาง", "example": "เงินอุดหนุนรายหัว", "note": "ต้องระบุเมื่อ isTransfer เป็น true"}
    ]
    
    # กำหนดสไตล์
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    description_font = Font(italic=True)
    description_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    example_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    note_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    # กำหนดขอบ
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # กำหนดการจัดตำแหน่ง
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # ฟังก์ชันสำหรับการจัดรูปแบบชีท
    def format_sheet(sheet, columns):
        # กำหนดความกว้างคอลัมน์
        sheet.column_dimensions['A'].width = 15  # คอลัมน์ header
        sheet.column_dimensions['B'].width = 30  # คอลัมน์ description
        sheet.column_dimensions['C'].width = 30  # คอลัมน์ example
        sheet.column_dimensions['D'].width = 40  # คอลัมน์ note
        
        # เขียนหัวข้อตาราง
        sheet['A1'] = "คอลัมน์"
        sheet['B1'] = "คำอธิบาย"
        sheet['C1'] = "ตัวอย่าง"
        sheet['D1'] = "ข้อควรระวัง"
        
        # จัดรูปแบบหัวข้อตาราง
        for col in range(1, 5):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # เขียนข้อมูลคอลัมน์
        for i, col_data in enumerate(columns, 2):
            sheet[f'A{i}'] = col_data["header"]
            sheet[f'B{i}'] = col_data["description"]
            sheet[f'C{i}'] = col_data["example"]
            sheet[f'D{i}'] = col_data["note"]
            
            # จัดรูปแบบเซลล์
            for col in range(1, 5):
                cell = sheet.cell(row=i, column=col)
                cell.border = thin_border
                
                if col == 1:  # คอลัมน์ header
                    cell.font = Font(bold=True)
                    cell.alignment = center_alignment
                elif col == 2:  # คอลัมน์ description
                    cell.font = description_font
                    cell.fill = description_fill
                    cell.alignment = left_alignment
                elif col == 3:  # คอลัมน์ example
                    cell.fill = example_fill
                    cell.alignment = left_alignment
                elif col == 4:  # คอลัมน์ note
                    cell.fill = note_fill
                    cell.alignment = left_alignment
        
        # เพิ่มตัวอย่างข้อมูล
        example_row = len(columns) + 3
        sheet[f'A{example_row}'] = "ตัวอย่างข้อมูล"
        sheet[f'A{example_row}'].font = Font(bold=True, size=14)
        sheet.merge_cells(f'A{example_row}:D{example_row}')
        
        # เขียนหัวข้อคอลัมน์สำหรับตัวอย่างข้อมูล
        for i, col_data in enumerate(columns, 1):
            sheet.cell(row=example_row + 1, column=i).value = col_data["header"]
            sheet.cell(row=example_row + 1, column=i).font = Font(bold=True)
            sheet.cell(row=example_row + 1, column=i).fill = header_fill
            sheet.cell(row=example_row + 1, column=i).alignment = center_alignment
            sheet.cell(row=example_row + 1, column=i).border = thin_border
        
        # เพิ่มตัวอย่างข้อมูลสำหรับ Projects
        if sheet.title == "Projects":
            example_data = [
                "PRJ001", "โครงการพัฒนาคุณภาพการศึกษา", "นายสมชาย ใจดี", "150000", 
                "academic", "2023-01-01", "2023-12-31", "active", 
                "โครงการเพื่อพัฒนาคุณภาพการศึกษา...", 
                '[{"category":"เงินอุดหนุนรายหัว","amount":100000},{"category":"เงินพัฒนาผู้เรียน","amount":50000}]'
            ]
        # เพิ่มตัวอย่างข้อมูลสำหรับ Transactions
        else:
            example_data = [
                "TRX001", "PRJ001", "2023-02-15", "จัดซื้อวัสดุการเรียน", "5000", 
                "เงินอุดหนุนรายหัว", "เบิกจ่ายตามใบเสร็จเลขที่ 12345", "false", "", "", "", "", ""
            ]
        
        # เขียนตัวอย่างข้อมูล
        for i, value in enumerate(example_data, 1):
            cell = sheet.cell(row=example_row + 2, column=i)
            cell.value = value
            cell.border = thin_border
            cell.alignment = left_alignment
    
    # จัดรูปแบบชีท Projects
    format_sheet(projects_sheet, projects_columns)
    
    # จัดรูปแบบชีท Transactions
    format_sheet(transactions_sheet, transactions_columns)
    
    # สร้างโฟลเดอร์ templates ถ้ายังไม่มี
    if not os.path.exists('templates'):
        os.makedirs('templates')
    
    # บันทึกไฟล์
    wb.save('templates/project_export_template.xlsx')
    print("สร้างไฟล์ Excel template เรียบร้อยแล้วที่ templates/project_export_template.xlsx")

if __name__ == "__main__":
    create_excel_template() 